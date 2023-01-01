import glob
import os
import re

import openpyxl as xl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


class MergingExcelFiles:

    @classmethod
    def creat_path_and_name_excel_file(cls) -> str:
        open_folder_path: str = input()
        name_of_file: str = "\\" + input() + ".xlsx"
        return open_folder_path + name_of_file

    @classmethod
    def creat_and_active_excel_file(cls) -> Worksheet:
        filepath: str = cls.creat_path_and_name_excel_file()
        wb: Worksheet = xl.Workbook() # creating a book object # создаем объект книги
        wb.save(filepath) # save it
        wb2: Worksheet = xl.load_workbook(filepath)
        ws2: Worksheet = wb2.active
        return wb2, ws2 

    @staticmethod
    def merge_excel_files(wb2: Worksheet, ws2: Worksheet, filepath: str) -> Worksheet:
        
        folder: None = os.chdir('C:\\Users\\vovch\\Desktop\\excel_n')  #folder with files
        xl_files : list = glob.glob('*.xlsx') # only files with extension .xlsx'

        for f in xl_files:
            wb1: Worksheet = xl.load_workbook(f)
            ws1: Worksheet = wb1.worksheets[0]
            mr: int = ws1.max_row
            mc: int = ws1.max_column
            wb2_max_row = ws2.max_row

            for i in range (1, mr + 1):
                for j in range (1, mc + 1):
                    c: Cell = ws1.cell(row = i, column = j)
                    ws2.cell(row = i+wb2_max_row, column = j).value = c.value
        wb2.save(str(filepath))

        return ws1




filepath = MergingExcelFiles.creat_path_and_name_excel_file()
wb2, ws2 = MergingExcelFiles.creat_and_active_excel_file(filepath)
MergingExcelFiles.merge_excel_files(wb2, ws2, filepath)

wb2, ws2 = MergingExcelFiles.creat_and_active_excel_file()
print(wb2, ws2)


filepath = creat_path_and_name_excel_file()
wb2, ws2 = creat_and_active_excel_file(filepath)
ws1 = merging_excel_files(wb2, ws2, filepath)
