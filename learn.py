# importing openpyxl module

import glob
import os

import openpyxl as xl

# book = openpyxl.open('C:\\Users\\vovch\\Desktop\\excel_n\\1.xlsx', read_only=True)
# sheet = book.active
# for row in range(1,sheet.max_row):
#     print(sheet[row])

# a = []
# for row in sheet.iter_rows():
#     print(row.value)




# filepath = 'C:\\Users\\vovch\\Desktop\\excel_n\\test.xlsx'
# wb = xl.Workbook()

# wb.save(filepath)



# # opening the source excel file
# filename ='C:\\Users\\vovch\\Desktop\\excel_n\\1.xlsx'
# wb1 = xl.load_workbook(filename)
# ws1 = wb1.worksheets[0]
  
# # opening the destination excel file 
# filename1 = 'C:\\Users\\vovch\\Desktop\\excel_n\\test.xlsx'
# wb2 = xl.load_workbook(filename1)
# ws2 = wb2.active
  
# # calculate total number of rows and 
# # columns in source excel file
# mr = ws1.max_row
# mc = ws1.max_column
  
# # copying the cell values from source 
# # excel file to destination excel file
# for i in range (1, mr + 1):
#     for j in range (1, mc + 1):
#         # reading cell value from source excel file
#         c = ws1.cell(row = i, column = j)
  
#         # writing the read value to destination excel file
#         ws2.cell(row = i, column = j).value = c.value
  
# # saving the destination excel file
# wb2.save(str(filename1))

# filepath = 'C:\\Users\\vovch\\Desktop\\excel_n\\test.xlsx'
# wb = xl.Workbook()

# wb.save(filepath)

filepath = 'C:\\Users\\vovch\\Desktop\\excel_n\\test.xlsx'
wb = xl.Workbook()

wb.save(filepath)

filename1 = 'C:\\Users\\vovch\\Desktop\\excel_n\\test.xlsx'
wb2 = xl.load_workbook(filename1)
ws2 = wb2.active


# folder = os.chdir('C:\\Users\\vovch\\Desktop\\excel_n')  #папка с файлами
# xl_files = glob.glob('*.xlsx')
# for f in xl_files:
#     wb1 = xl.load_workbook(f)
#     ws1 = wb1.worksheets[0]
#     mr = ws1.max_row
#     mc = ws1.max_column
#     wb2_max_row = ws2.max_row

#     for i in range (1, mr + 1):
#         for j in range (1, mc + 1):
#             c = ws1.cell(row = i, column = j)
    
#             # writing the read value to destination excel file
#             ws2.cell(row = i+wb2_max_row, column = j).value = c.value

  
# saving the destination excel file
wb2.save(str(filename1))