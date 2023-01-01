import glob
import os
import re

import openpyxl as xl
from openpyxl.cell.cell import Cell
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

FIND_DECLARATION = 'Декларация на страхование грузов №'

def creat_path_and_name_excel_file() -> str:
    open_folder_path: str = input()
    name_of_file: str = "\\" + input() + ".xlsx"
    return open_folder_path + name_of_file


def creat_and_active_excel_file(filepath: str) -> Worksheet:
    wb: Worksheet = xl.Workbook() # creating a book object # создаем объект книги
    wb.save(filepath) # save it
    wb2: Worksheet = xl.load_workbook(filepath)
    ws2: Worksheet = wb2.active
    return wb2, ws2 


def merging_excel_files(wb2: Worksheet, ws2: Worksheet, filepath: str) -> Worksheet:
    
    folder: None = os.chdir('C:\\Users\\vovch\\Desktop\\excel_n')  #folder with files
    xl_files : list = glob.glob('*.xlsx') # only files with extension .xlsx'

    for f in xl_files:
        wb1: Worksheet = xl.load_workbook(f)
        ws1: Worksheet = wb1.worksheets[0]
        mr: int = ws1.max_row
        mc: int = ws1.max_column
        wb2_max_row = ws2.max_row

        for i in range (1, mr + 1):
            for j in range (1, mc + 1):
                c: Cell = ws1.cell(row = i, column = j)
                ws2.cell(row = i+wb2_max_row, column = j).value = c.value
    wb2.save(str(filepath))

    return ws1

    
# Может понадобится!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!
# from openpyxl import Workbook,load_workbook

# wbSearch = Workbook()
# wbSearch = load_workbook("search.xlsx")
# wsSearch = wbSearch.active

# wbResult = Workbook()
# wsResult = wbResult.active
# resultRow = 1

# lookFor = 'Блендер'

# for i in range(1,21):
#    value=wsSearch.cell(row=i, column=1).value
#    if value == lookFor:
#        wsResult.cell(row=resultRow, column=1).value=value

# wbResult.save("result.xlsx")

def search_in_excel_files(ws1: Worksheet) -> None:
    row_max = ws2.max_row  # Получаем количество столбцов
    column_max = ws2.max_column  # Получаем количество строк
    row_min = 1 #Переменная, отвечающая за номер строки
    column_min = 1 #Переменная, отвечающая за номер столбца
    search_text = "Декларация"
    while column_min <= column_max:
        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)
    
            word_column = get_column_letter(column_min)
            word_column = str(word_column)
            word_cell = word_column + row_min_min
    
            data_from_cell = ws1[word_cell].value
            data_from_cell = str(data_from_cell)
            print(data_from_cell)
            regular = search_text
            result = re.findall(regular, data_from_cell)
            if len(result) > 0:
                print('Нашли в ячейке:', word_cell)
            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1
        column_min = column_min + 1

filepath = creat_path_and_name_excel_file()
wb2, ws2 = creat_and_active_excel_file(filepath)
ws1 = merging_excel_files(wb2, ws2, filepath)
search_in_excel_files(ws1)




  
